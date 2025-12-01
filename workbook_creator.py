"""Модуль для создания новых выгрузок Excel."""
import calendar
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path


class WorkbookCreator:
    def __init__(self):
        pass

    def create_workbook(self, data, activity_name, output_dir):
        total:float = 0
        output_dir = Path(output_dir) 
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "RUB"
        column_widths = []

        output_dir.mkdir(parents=True, exist_ok=True)

        headers = [
            'Проект', 'Категория', 'Наименование услуги',
            'Тариф', 'Расчетный тариф', 'Количество',
            'Единица измерения', 'Стоимость'
        ]
        sheet.append(headers)

        for row in data:
            sheet.append(row)
            total += float(row[-1])
            
            for i, cell in enumerate(row):
                cell_str = str(cell)
                if len(column_widths) > i:
                    if len(cell_str) > column_widths[i]:
                            column_widths[i] = len(cell_str)
                else:
                    column_widths += [len(cell_str)]
        
        for i, column_width in enumerate(column_widths,1):  
            sheet.column_dimensions[get_column_letter(i)].width = column_width
        
        for cell in sheet["1:1"]:
            cell.font = cell.font.copy(bold=True)

        sheet.append(["Итого", total])
        month_name = calendar.month_name[datetime.now().month - 1 if datetime.now().month > 1 else 12]

        filename = f"Отчет_{activity_name}_{month_name}.xlsx"
        file_path = output_dir / filename

        workbook.save(file_path)
        return str(file_path)

